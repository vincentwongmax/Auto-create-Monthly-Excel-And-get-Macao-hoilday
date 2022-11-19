import datetime
import urllib.request
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import  PatternFill , Border, Side
from datetime import datetime
import tkinter as tk
from tkinter import simpledialog
from tkinter import messagebox
import pandas as pd


urllib.request.urlretrieve("https://www.gov.mo/zh-hant/public-holidays/ical-timestamp/", "hoilday.ics")  ## 使用urllib.request下載ics 檔案

wb = load_workbook('origin.xlsx') 
ws =wb.active

# today = datetime.datetime.today()     ## 今天的日期
# year = today.year   ##從年月日中取出年份
# #month = today.month   ##從年月日中取出月份
# month = 2 ##從年月日中取出月份

ROOT = tk.Tk()
ROOT.withdraw()
year = simpledialog.askstring(title="Please input Year",prompt="Enter year:")
month = simpledialog.askstring(title="Please input Month",prompt="Enter Month:")

# year = (input("Enter year:")) ##從年月日中取出月份
# month = (input("Enter Month:")) ##從年月日中取出月份
year_int = int(year)
month_int = int(month)

if len(month) == 1 :
    month = '0' + str(month) 

date_string = "{}-{}-01 10:10:10".format(year,month)

if month_int >12 or month_int <1 :
    messagebox.showerror("Error", "The input is incorrect, the corresponding month cannot be found")

today = datetime.fromisoformat(date_string)


# print(month)
# print(result)
# result.insert(len(result), 'c')
#result.append() 
# print(result)
result = []
j=0
for i in range (1,40):    ## 自訂年月日的格式, 存入result 的陣列

    if len(str(i)) == 1 :
        j = '0' + str(i) 
    else:
        j = i

    result.insert(len(result), str(year) + "/" + str(month) + "/" + str(j))
    
   # date_for_weekend.insert(len(date_for_weekend), str(year) + "-" + str(month) + "-" + str(i))

    #result = str(year) + "/" + str(month) + "/" + str(i)
    # print(year,"/",month,"/",i,sep="")   ## print 使用的方法

date_for_weekend = []

if month_int == 2:  #28日
    for i in ["01","02","03","04","05","06","07","08","09","10","11","12","13","14","15","16","17","18","19","20","21","22","23","24","25","26","27","28"]:
        date_for_weekend.insert(len(date_for_weekend),datetime.fromisoformat("{}-{}-{} 10:10:10".format(year,month,str(i))))
elif month_int == 1 or month_int == 3 or month_int == 5 or month_int == 7 or month_int == 8 or month_int == 10 or month_int == 12:    #31日  
    for i in ["01","02","03","04","05","06","07","08","09","10","11","12","13","14","15","16","17","18","19","20","21","22","23","24","25","26","27","28","29","30","31"]:
        date_for_weekend.insert(len(date_for_weekend),datetime.fromisoformat("{}-{}-{} 10:10:10".format(year,month,str(i))))
else :  #30日  其他的都剩下30日
    for i in ["01","02","03","04","05","06","07","08","09","10","11","12","13","14","15","16","17","18","19","20","21","22","23","24","25","26","27","28","29","30"]:
        date_for_weekend.insert(len(date_for_weekend),datetime.fromisoformat("{}-{}-{} 10:10:10".format(year,month,str(i))))

print("有多少個陣列",len(date_for_weekend))


if month_int == 2:  # 2月
    ranges = 31   #原應該由28開始，但EXCEL 儲存格由A3開始，所以加3

    ws["A33"].style = 'Normal'  #no 31 
    ws["B33"].style = 'Normal'
    ws["C33"].style = 'Normal'
    ws["D33"].style = 'Normal'
    ws["E33"].style = 'Normal'
    ws["F33"].style = 'Normal'
    ws["G33"].style = 'Normal'

    ws["A32"].style = 'Normal'   #no 30
    ws["B32"].style = 'Normal'
    ws["C32"].style = 'Normal'
    ws["D32"].style = 'Normal'
    ws["E32"].style = 'Normal'
    ws["F32"].style = 'Normal'
    ws["G32"].style = 'Normal'

    ws["A31"].style = 'Normal'   # no 29
    ws["B31"].style = 'Normal'
    ws["C31"].style = 'Normal'
    ws["D31"].style = 'Normal'
    ws["E31"].style = 'Normal'
    ws["F31"].style = 'Normal'
    ws["G31"].style = 'Normal'

elif month_int == 1 or month_int == 3 or month_int == 5 or month_int == 7 or month_int == 8 or month_int == 10 or month_int == 12:   ##31 日的月份
    ranges = 34    #原應該由31開始，但EXCEL 儲存格由A3開始，所以加3
else:  # 其他都是30號的月份
    ranges = 33   
    #原應該由30開始，但EXCEL 儲存格由A3開始，所以加3
  
    ws["A33"].style = 'Normal'  #no 31
    ws["B33"].style = 'Normal'
    ws["C33"].style = 'Normal'
    ws["D33"].style = 'Normal'
    ws["E33"].style = 'Normal'
    ws["F33"].style = 'Normal'
    ws["G33"].style = 'Normal'
    
############
    # side = Side(     ###边框颜色
    #     style="thin",  # 边框样式，可选dashDot、dashDotDot、dashed、dotted、double、hair、medium、mediumDashDot、mediumDashDotDot、mediumDashed、slantDashDot、thick、thin
    #     color="FFFFFF",  # 边框颜色，16进制rgb表示
    # )


    # ws["A33"].border = Border(
    #     #top=side,     # 上
    #     bottom=side,  # 下
    #     left=side,    # 左
    #     right=side,   # 右
    # )

    # ws["B33"].border = Border(
    #     #top=side,     # 上
    #     bottom=side,  # 下
    #     left=side,    # 左
    #     right=side,   # 右
    # )

    # ws["C33"].border = Border(
    #     #top=side,     # 上
    #     bottom=side,  # 下
    #     left=side,    # 左
    #     right=side,   # 右
    # )

    # ws["D33"].border = Border(
    #     #top=side,     # 上
    #     bottom=side,  # 下
    #     left=side,    # 左
    #     right=side,   # 右
    # )

    # ws["E33"].border = Border(
    #     #top=side,     # 上
    #     bottom=side,  # 下
    #     left=side,    # 左
    #     right=side,   # 右
    # )

    # ws["F33"].border = Border(
    #     #top=side,     # 上
    #     bottom=side,  # 下
    #     left=side,    # 左
    #     right=side,   # 右
    # )

    # ws["G33"].border = Border(
    #     #top=side,     # 上
    #     bottom=side,  # 下
    #     left=side,    # 左
    #     right=side,   # 右
    # )
##################


for j in range(3,ranges):
    ws['A{}'.format(j)].value = result[j-3]     ### 更改excel 儲存格的值
    # print(ws['A{}'.format(i)].value) 
    # print(ws['A1'].value)
    ws['B{}'.format(j)].value = '=CHOOSE(WEEKDAY(A{},1),"Sunday","Monday","Tuesday","Wednesday","Thursday","Friday","Saturday")'.format(j)   ### 更改excel 儲存格的值
  #  print ("Excel Value", ws['A{}'.format(j)].value,"   result   ",result[j-3] ,"    weekendday    ",date_for_weekend[j-3] , "weekend" ,date_for_weekend[j-3].strftime("%w"),"A",j)

    if int(date_for_weekend[j-3].strftime("%w")) == 6 or int(date_for_weekend[j-3].strftime("%w")) == 0:  ##找出星期六日儲存格作填色
        #if j+2 < ranges :
        # print("日期" , date_for_weekend[j-3],"星期" , date_for_weekend[j-3].strftime("%w")) 
        # print("A幾",j)
        ws['A{}'.format(j)].fill = PatternFill("solid", fgColor="FFFF99")  ## 相對應的儲存格作填色
        ws['B{}'.format(j)].fill = PatternFill("solid", fgColor="FFFF99")
        ws['C{}'.format(j)].fill = PatternFill("solid", fgColor="FFFF99")
        ws['D{}'.format(j)].fill = PatternFill("solid", fgColor="FFFF99")
        ws['E{}'.format(j)].fill = PatternFill("solid", fgColor="FFFF99")
        ws['F{}'.format(j)].fill = PatternFill("solid", fgColor="FFFF99")
        ws['G{}'.format(j)].fill = PatternFill("solid", fgColor="FFFF99")

from icalendar import Calendar, Event  
from pytz import UTC # timezone
g = open('hoilday.ics','rb')   ## 讀入hoilday.ics 日曆
gcal = Calendar.from_ical(g.read())

Have_This_Year_Month = 0 

for component in gcal.walk():
    if component.name == "VEVENT":
        # print(component.get('summary'))
        # print(component.get('dtstart'))
        # print((component.get('dtend')))
        # print(component.get('dtstamp'))
        date = component.decoded('dtend')     ## 找出重要日子的最後日期，因為日期是 UTC ，所以直接取 dtend 時間會對上 +8 時區
       # print(str(date.year),str(date.month))
        if date.year == year_int and date.month == month_int:           ## 找出今年 和今個月
            ws['A{}'.format(date.day + 2 )].fill = PatternFill("solid", fgColor="FCE4D6")  ## 相對應的儲存格作填色
            ws['B{}'.format(date.day + 2 )].fill = PatternFill("solid", fgColor="FCE4D6")
            ws['C{}'.format(date.day + 2 )].fill = PatternFill("solid", fgColor="FCE4D6")
            ws['D{}'.format(date.day + 2 )].fill = PatternFill("solid", fgColor="FCE4D6")
            ws['E{}'.format(date.day + 2 )].fill = PatternFill("solid", fgColor="FCE4D6")
            ws['F{}'.format(date.day + 2 )].fill = PatternFill("solid", fgColor="FCE4D6")
            ws['G{}'.format(date.day + 2 )].fill = PatternFill("solid", fgColor="FCE4D6")
        
        if date.year == year_int :                      ## 判斷用戶輸入的數字是否有這個年份
            Have_This_Year_Month = 1
        #print(type(date))

g.close()

bigmonth = today.strftime("%B")


if Have_This_Year_Month == 1 and month_int <= 12 and month_int >= 1 :
    wb.save('Daily Report of CAM 4F 5F-{}.xlsx'.format(bigmonth))
else :
    messagebox.showerror("Did NOT HAVA DATA", "The input is incorrect, the corresponding year or month cannot be found")

