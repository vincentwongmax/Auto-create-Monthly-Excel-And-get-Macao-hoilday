import datetime
import urllib.request
from openpyxl import  load_workbook
from openpyxl.styles import  PatternFill  
from datetime import datetime
import tkinter as tk
from tkinter import simpledialog
from tkinter import messagebox
from icalendar import Calendar  

urllib.request.urlretrieve("https://www.gov.mo/zh-hant/public-holidays/ical-timestamp/", "hoilday.ics") 
wb = load_workbook('origin.xlsx') 
ws =wb.active
ROOT = tk.Tk()
ROOT.withdraw()
year = simpledialog.askstring(title="Please input Year",prompt="Enter year:")
month = simpledialog.askstring(title="Please input Month",prompt="Enter Month:")
year_int = int(year)
month_int = int(month)

if len(month) == 1 :
    month = '0' + str(month) 
date_string = "{}-{}-01 10:10:10".format(year,month)
if month_int >12 or month_int <1 :
    messagebox.showerror("Error", "The input is incorrect, the corresponding month cannot be found")
today = datetime.fromisoformat(date_string)

result = []
for i in range (1,40):   
    if len(str(i)) == 1 :
        j = '0' + str(i) 
    else:
        j = i
    result.insert(len(result), str(year) + "/" + str(month) + "/" + str(j))
date_for_weekend = []

if month_int == 2:  
    for i in ["01","02","03","04","05","06","07","08","09","10","11","12","13","14","15","16","17","18","19","20","21","22","23","24","25","26","27","28"]:
        date_for_weekend.insert(len(date_for_weekend),datetime.fromisoformat("{}-{}-{} 10:10:10".format(year,month,str(i))))
elif month_int == 1 or month_int == 3 or month_int == 5 or month_int == 7 or month_int == 8 or month_int == 10 or month_int == 12:   
    for i in ["01","02","03","04","05","06","07","08","09","10","11","12","13","14","15","16","17","18","19","20","21","22","23","24","25","26","27","28","29","30","31"]:
        date_for_weekend.insert(len(date_for_weekend),datetime.fromisoformat("{}-{}-{} 10:10:10".format(year,month,str(i))))
else :  
    for i in ["01","02","03","04","05","06","07","08","09","10","11","12","13","14","15","16","17","18","19","20","21","22","23","24","25","26","27","28","29","30"]:
        date_for_weekend.insert(len(date_for_weekend),datetime.fromisoformat("{}-{}-{} 10:10:10".format(year,month,str(i))))

if month_int == 2:  
    ranges = 31   
    ws["A33"].style = 'Normal' 
    ws["B33"].style = 'Normal'
    ws["C33"].style = 'Normal'
    ws["D33"].style = 'Normal'
    ws["E33"].style = 'Normal'
    ws["F33"].style = 'Normal'
    ws["G33"].style = 'Normal'
    ws["A32"].style = 'Normal'   
    ws["B32"].style = 'Normal'
    ws["C32"].style = 'Normal'
    ws["D32"].style = 'Normal'
    ws["E32"].style = 'Normal'
    ws["F32"].style = 'Normal'
    ws["G32"].style = 'Normal'
    ws["A31"].style = 'Normal'   
    ws["B31"].style = 'Normal'
    ws["C31"].style = 'Normal'
    ws["D31"].style = 'Normal'
    ws["E31"].style = 'Normal'
    ws["F31"].style = 'Normal'
    ws["G31"].style = 'Normal'
elif month_int == 1 or month_int == 3 or month_int == 5 or month_int == 7 or month_int == 8 or month_int == 10 or month_int == 12:  
    ranges = 34   
else: 
    ranges = 33   
    ws["A33"].style = 'Normal' 
    ws["B33"].style = 'Normal'
    ws["C33"].style = 'Normal'
    ws["D33"].style = 'Normal'
    ws["E33"].style = 'Normal'
    ws["F33"].style = 'Normal'
    ws["G33"].style = 'Normal'
    
for j in range(3,ranges):
    ws['A{}'.format(j)].value = result[j-3]     
    ws['B{}'.format(j)].value = '=CHOOSE(WEEKDAY(A{},1),"Sunday","Monday","Tuesday","Wednesday","Thursday","Friday","Saturday")'.format(j)   
  
    if int(date_for_weekend[j-3].strftime("%w")) == 6 or int(date_for_weekend[j-3].strftime("%w")) == 0:  
        ws['A{}'.format(j)].fill = PatternFill("solid", fgColor="FFFF99")  
        ws['B{}'.format(j)].fill = PatternFill("solid", fgColor="FFFF99")
        ws['C{}'.format(j)].fill = PatternFill("solid", fgColor="FFFF99")
        ws['D{}'.format(j)].fill = PatternFill("solid", fgColor="FFFF99")
        ws['E{}'.format(j)].fill = PatternFill("solid", fgColor="FFFF99")
        ws['F{}'.format(j)].fill = PatternFill("solid", fgColor="FFFF99")
        ws['G{}'.format(j)].fill = PatternFill("solid", fgColor="FFFF99")

g = open('hoilday.ics','rb')   
gcal = Calendar.from_ical(g.read())
Have_This_Year_Month = 0 
for component in gcal.walk():
    if component.name == "VEVENT":
        date = component.decoded('dtend')     
        if date.year == year_int and date.month == month_int:          
            ws['A{}'.format(date.day + 2 )].fill = PatternFill("solid", fgColor="FCE4D6")  
            ws['B{}'.format(date.day + 2 )].fill = PatternFill("solid", fgColor="FCE4D6")
            ws['C{}'.format(date.day + 2 )].fill = PatternFill("solid", fgColor="FCE4D6")
            ws['D{}'.format(date.day + 2 )].fill = PatternFill("solid", fgColor="FCE4D6")
            ws['E{}'.format(date.day + 2 )].fill = PatternFill("solid", fgColor="FCE4D6")
            ws['F{}'.format(date.day + 2 )].fill = PatternFill("solid", fgColor="FCE4D6")
            ws['G{}'.format(date.day + 2 )].fill = PatternFill("solid", fgColor="FCE4D6")
        if date.year == year_int :                     
            Have_This_Year_Month = 1
g.close()
bigmonth = today.strftime("%B")
if Have_This_Year_Month == 1 and month_int <= 12 and month_int >= 1 :
    wb.save('Daily Report of CAM 4F 5F-{}.xlsx'.format(bigmonth))
else :
    messagebox.showerror("Did NOT HAVA DATA", "The input is incorrect, the corresponding year or month cannot be found")

