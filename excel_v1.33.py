import datetime
import urllib.request
from openpyxl import  load_workbook
from openpyxl.styles import  PatternFill  
from datetime import datetime
import tkinter as tk
from tkinter import simpledialog
from tkinter import messagebox
from icalendar import Calendar  
import calendar as calendarisleap
import os,time
import configparser

config=configparser.ConfigParser()
config.read('conf.ini')
conf_day_diff = int(config.get('conf','day_diff'))
conf_weekend_color = config.get('conf','weekend_color')
conf_hoilday_color = config.get('conf','hoilday_color')
conf_save_excel_name = config.get('conf','save_excel_name')
conf_year_month_day_who_first = config.get('conf','year_month_day_who_first')
conf_year_month_day_who_second = config.get('conf','year_month_day_who_second')
conf_year_month_day_who_third = config.get('conf','year_month_day_who_third')
conf_excel_Formula_mode = config.get('conf','excel_Formula_mode')
conf_excel_output_year_insavename = config.get('conf','excel_output_year_insavename')
conf_excel_output__sort_insavename = config.get('conf','excel_output__sort_insavename')


conf_muti_mode_on = config.get('muti_mode','muti_mode_on')
conf_export_year = config.get('muti_mode','export_year')
conf_export_start_month= int(config.get('muti_mode','export_start_month'))
conf_export_end_month = int(config.get('muti_mode','export_end_month')) + 1 

if conf_muti_mode_on != 'true' :
    conf_export_start_month = 1
    conf_export_end_month = 2

for loop_how_time in range (conf_export_start_month,conf_export_end_month):
    def demo(day1, day2):  
        time_array1 = time.strptime(day1, "%Y-%m-%d")     
        timestamp_day1 = int(time.mktime(time_array1))   
        time_array2 = time.strptime(day2, "%Y-%m-%d")     
        timestamp_day2 = int(time.mktime(time_array2))   
        result = (timestamp_day2 - timestamp_day1) // 60 // 60 // 24   
        return result

    filepath = "hoilday.ics"
        
    if os.path.isfile(filepath):
        modifiedTime = time.localtime(os.stat(filepath).st_mtime)  
        createdTime = time.localtime(os.stat(filepath).st_ctime)   
        mTimes = time.strftime('%Y-%m-%d', modifiedTime)   
        cTimes = time.strftime('%Y-%m-%d', createdTime)   
        today = datetime.today().strftime("%Y-%m-%d")  
        day_diff = demo(mTimes, today)
        print("兩個日期相隔的天數：{} ".format(day_diff))

        if int(day_diff) < conf_day_diff :
            print("hoilday.ics檔案存在, 版本是當天的，不會從官方下載")
        else:
            print("hoilday.ics檔案存在, 但版本過舊，立即下載最新的檔案")
            urllib.request.urlretrieve("https://www.gov.mo/zh-hant/public-holidays/ical-timestamp/", "hoilday.ics") 
    else:
        print("hoilday.ics檔案不存在,立即下載最新的澳門假期")
        urllib.request.urlretrieve("https://www.gov.mo/zh-hant/public-holidays/ical-timestamp/", "hoilday.ics") 

    wb = load_workbook('origin.xlsx') 
    ws =wb.active
    ROOT = tk.Tk()
    ROOT.withdraw()
    
    if conf_muti_mode_on == 'true' :
        year = conf_export_year
        month = str(loop_how_time)
    else:
        year = simpledialog.askstring(title="Please input Year",prompt="Enter year:")
        month = simpledialog.askstring(title="Please input Month",prompt="Enter Month:")
    year_int = int(year)
    month_int = int(month)

    if len(month) == 1 :
        month = '0' + str(month) 
    date_string = "{}-{}-01 10:10:10".format(year,month)
    if month_int >12 or month_int <1 :
        messagebox.showerror("Error", "The input is incorrect, the corresponding month cannot be found {}_{}".format(year_int,month_int))
    today = datetime.fromisoformat(date_string)

    result = []
    for i in range (1,40):   
        if len(str(i)) == 1 :
            day = '0' + str(i) 
        else:
            day = i
        result.insert(len(result), str(globals()[conf_year_month_day_who_first]) + "/" + str(globals()[conf_year_month_day_who_second]) + "/" + str(globals()[conf_year_month_day_who_third]))
    date_for_weekend = []

    if month_int == 2:  
        if calendarisleap.isleap(year_int) == True :
            for i in ["01","02","03","04","05","06","07","08","09","10","11","12","13","14","15","16","17","18","19","20","21","22","23","24","25","26","27","28","29"]:
                date_for_weekend.insert(len(date_for_weekend),datetime.fromisoformat("{}-{}-{} 10:10:10".format(year,month,str(i))))
        else:
            for i in ["01","02","03","04","05","06","07","08","09","10","11","12","13","14","15","16","17","18","19","20","21","22","23","24","25","26","27","28"]:
                date_for_weekend.insert(len(date_for_weekend),datetime.fromisoformat("{}-{}-{} 10:10:10".format(year,month,str(i))))
    elif month_int == 1 or month_int == 3 or month_int == 5 or month_int == 7 or month_int == 8 or month_int == 10 or month_int == 12:   
        for i in ["01","02","03","04","05","06","07","08","09","10","11","12","13","14","15","16","17","18","19","20","21","22","23","24","25","26","27","28","29","30","31"]:
            date_for_weekend.insert(len(date_for_weekend),datetime.fromisoformat("{}-{}-{} 10:10:10".format(year,month,str(i))))
    else :  
        for i in ["01","02","03","04","05","06","07","08","09","10","11","12","13","14","15","16","17","18","19","20","21","22","23","24","25","26","27","28","29","30"]:
            date_for_weekend.insert(len(date_for_weekend),datetime.fromisoformat("{}-{}-{} 10:10:10".format(year,month,str(i))))

    if month_int == 2:  
        ranges = 31   
        ws["A33"].style = 'Normal' 
        ws["B33"].style = 'Normal'
        ws["C33"].style = 'Normal'
        ws["D33"].style = 'Normal'
        ws["E33"].style = 'Normal'
        ws["F33"].style = 'Normal'
        ws["G33"].style = 'Normal'
        ws["A32"].style = 'Normal'   
        ws["B32"].style = 'Normal'
        ws["C32"].style = 'Normal'
        ws["D32"].style = 'Normal'
        ws["E32"].style = 'Normal'
        ws["F32"].style = 'Normal'
        ws["G32"].style = 'Normal'

        if calendarisleap.isleap(year_int) == False :
            ws["A31"].style = 'Normal'   
            ws["B31"].style = 'Normal'
            ws["C31"].style = 'Normal'
            ws["D31"].style = 'Normal'
            ws["E31"].style = 'Normal'
            ws["F31"].style = 'Normal'
            ws["G31"].style = 'Normal'
    elif month_int == 1 or month_int == 3 or month_int == 5 or month_int == 7 or month_int == 8 or month_int == 10 or month_int == 12:  
        ranges = 34   
    else: 
        ranges = 33   
        ws["A33"].style = 'Normal' 
        ws["B33"].style = 'Normal'
        ws["C33"].style = 'Normal'
        ws["D33"].style = 'Normal'
        ws["E33"].style = 'Normal'
        ws["F33"].style = 'Normal'
        ws["G33"].style = 'Normal'
        
    for j in range(3,ranges):
        ws['A{}'.format(j)].value = result[j-3]     


        if conf_excel_Formula_mode == 'false' :
            ws['B{}'.format(j)].value = date_for_weekend[j-3].strftime("%A")
        else:
            ws['B{}'.format(j)].value = '=CHOOSE(WEEKDAY(A{},1),"Sunday","Monday","Tuesday","Wednesday","Thursday","Friday","Saturday")'.format(j)   
    
        if int(date_for_weekend[j-3].strftime("%w")) == 6 or int(date_for_weekend[j-3].strftime("%w")) == 0:  
            ws['A{}'.format(j)].fill = PatternFill("solid", fgColor="{}".format(conf_weekend_color))  
            ws['B{}'.format(j)].fill = PatternFill("solid", fgColor="{}".format(conf_weekend_color))
            ws['C{}'.format(j)].fill = PatternFill("solid", fgColor="{}".format(conf_weekend_color))
            ws['D{}'.format(j)].fill = PatternFill("solid", fgColor="{}".format(conf_weekend_color))
            ws['E{}'.format(j)].fill = PatternFill("solid", fgColor="{}".format(conf_weekend_color))
            ws['F{}'.format(j)].fill = PatternFill("solid", fgColor="{}".format(conf_weekend_color))
            ws['G{}'.format(j)].fill = PatternFill("solid", fgColor="{}".format(conf_weekend_color))

    g= open('hoilday.ics','rb')   
    gcal = Calendar.from_ical(g.read())
    Have_This_Year_Month = 0 
    for component in gcal.walk():
        if component.name == "VEVENT":
            date = component.decoded('dtend')     
            if date.year == year_int and date.month == month_int:
                ws['A{}'.format(date.day + 2 )].fill = PatternFill("solid", fgColor="{}".format(conf_hoilday_color))  
                ws['B{}'.format(date.day + 2 )].fill = PatternFill("solid", fgColor="{}".format(conf_hoilday_color))
                ws['C{}'.format(date.day + 2 )].fill = PatternFill("solid", fgColor="{}".format(conf_hoilday_color))
                ws['D{}'.format(date.day + 2 )].fill = PatternFill("solid", fgColor="{}".format(conf_hoilday_color))
                ws['E{}'.format(date.day + 2 )].fill = PatternFill("solid", fgColor="{}".format(conf_hoilday_color))
                ws['F{}'.format(date.day + 2 )].fill = PatternFill("solid", fgColor="{}".format(conf_hoilday_color))
                ws['G{}'.format(date.day + 2 )].fill = PatternFill("solid", fgColor="{}".format(conf_hoilday_color))
            if date.year == year_int :                     
                Have_This_Year_Month = 1
    g.close()
    bigmonth = today.strftime("%B")
    # map month integer (1-12) to sheet short names
    monthlyname = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']
    # set worksheet title according to the selected month
    try:
        ws.title = monthlyname[month_int - 1]
    except Exception as e:
        # If renaming fails (duplicate name or invalid), fall back to a safe title
        try:
            ws.title = f"{monthlyname[month_int - 1]}_{month_int}"
        except Exception:
            # last resort: leave original title
            pass
    #### 
    if conf_excel_output__sort_insavename == 'true' :
        output_sort = str(loop_how_time) + '. '
    else :
        output_sort = "" 
    ####
    if conf_excel_output_year_insavename == 'true' :
        output_year = str(year_int) + '-'
    else :
        output_year = ""
    ####

    if Have_This_Year_Month == 1 and month_int <= 12 and month_int >= 1 :
        wb.save('{}{}-{}{}.xlsx'.format(output_sort,conf_save_excel_name,output_year,bigmonth))
    else :
        messagebox.showerror("DID NOT HAVA DATA", "The input is incorrect, the corresponding year or month cannot be found {}_{}".format(year_int,month_int))

